#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
從 Excel 檔案解析韻母表並生成 JSON 資料
支援：
- 合併儲存格的正確處理
- 部分底線格式的自動識別（Rich Text）
- 底部邊框的識別
"""

import openpyxl
from openpyxl.styles import Border
import json
import os
import sys

def is_cell_merged(ws, row, col):
    """檢查儲存格是否為合併儲存格的一部分"""
    cell = ws.cell(row, col)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None

def get_merged_cell_value(ws, row, col):
    """取得合併儲存格的值（從左上角儲存格取得）"""
    merged_range = is_cell_merged(ws, row, col)
    if merged_range:
        # 取得合併區域的左上角儲存格
        # bounds 格式：(min_col, min_row, max_col, max_row)
        min_col, min_row, max_col, max_row = merged_range.bounds
        return ws.cell(min_row, min_col).value, (max_row - min_row + 1)
    return ws.cell(row, col).value, 1

def has_underline(cell):
    """檢查儲存格是否有底線格式"""
    if cell.font and cell.font.underline:
        return True
    return False

def has_bottom_border(cell):
    """檢查儲存格是否有底部粗框線"""
    if cell.border and cell.border.bottom:
        # 檢查是否為粗線或雙線
        style = cell.border.bottom.style
        if style in ['medium', 'thick', 'double']:
            return True
    return False

def wrap_underline(text):
    """將文本包裹在 <u> 標籤中"""
    if text:
        return f"<u>{text}</u>"
    return text

def extract_rich_text_with_underline(cell):
    """
    提取儲存格的 Rich Text，保留底線格式
    支援部分文字有底線的情況
    """
    value = cell.value

    # 如果是純文字
    if isinstance(value, str):
        # 檢查整個儲存格是否有底線
        if has_underline(cell):
            return wrap_underline(value)
        return value

    # 如果是 Rich Text（InlineFont 格式）
    # openpyxl 3.x 版本中，Rich Text 儲存為 list 或特殊物件
    if hasattr(value, '__iter__') and not isinstance(value, str):
        result = []
        for item in value:
            # 檢查是否有 text 和 font 屬性
            if hasattr(item, 'text'):
                text = item.text
                # 檢查是否有底線
                has_u = False
                if hasattr(item, 'font') and item.font:
                    # 檢查多種底線屬性
                    if hasattr(item.font, 'underline') and item.font.underline:
                        has_u = True
                    elif hasattr(item.font, 'u') and item.font.u:
                        has_u = True

                if has_u:
                    result.append(wrap_underline(text))
                else:
                    result.append(text)
            elif isinstance(item, str):
                result.append(item)

        return ''.join(result) if result else str(value)

    # 其他情況，直接返回字串
    return str(value) if value else ''

def parse_excel_rhyme_table(file_path):
    """解析 Excel 韻母表"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    result = []

    # 記錄每個方言點欄位的跨行狀態
    dialect_merge_state = {}  # {col_index: {'value': '...', 'remaining': N}}
    letter_merge_state = None  # 韻母字欄位的跨行狀態
    phonetic_merge_state = None  # 擬音欄位的跨行狀態

    # 從第3行開始讀取（前2行是表頭）
    for row_idx in range(3, ws.max_row + 1):
        row_data = {
            'letter': '',
            'examples': '',
            'phonetic': '',
            'dialectValues': [],
            'rowSpan': 1,
            'phoneticRowSpan': 1,
            'hasBorder': False
        }

        # Column A: 韻母字（平話字）
        letter_cell = ws.cell(row_idx, 1)
        letter_merged_range = is_cell_merged(ws, row_idx, 1)

        if letter_merged_range:
            # 儲存格在合併區域內
            min_col, min_row, max_col, max_row = letter_merged_range.bounds
            if row_idx == min_row:
                # 這是合併區域的第一行
                letter_value = letter_cell.value
                letter_rowspan = max_row - min_row + 1
                row_data['letter'] = str(letter_value) if letter_value else ''
                row_data['rowSpan'] = letter_rowspan
                if letter_rowspan > 1:
                    letter_merge_state = {
                        'value': str(letter_value) if letter_value else '',
                        'remaining': letter_rowspan - 1
                    }
            else:
                # 這是合併區域的後續行
                if letter_merge_state and letter_merge_state['remaining'] > 0:
                    row_data['letter'] = letter_merge_state['value']
                    row_data['rowSpan'] = 0
                    letter_merge_state['remaining'] -= 1
                else:
                    # 應該不會到這裡，但以防萬一
                    letter_value = ws.cell(min_row, min_col).value
                    row_data['letter'] = str(letter_value) if letter_value else ''
                    row_data['rowSpan'] = 0
        else:
            # 非合併儲存格
            letter_value = letter_cell.value
            row_data['letter'] = str(letter_value) if letter_value else ''
            row_data['rowSpan'] = 1

        # Column B: 例字（支援 Rich Text 底線）
        examples_cell = ws.cell(row_idx, 2)
        row_data['examples'] = extract_rich_text_with_underline(examples_cell)

        # Column C: 擬音（支援合併儲存格）
        phonetic_cell = ws.cell(row_idx, 3)
        phonetic_merged_range = is_cell_merged(ws, row_idx, 3)

        if phonetic_merged_range:
            # 儲存格在合併區域內
            min_col, min_row, max_col, max_row = phonetic_merged_range.bounds
            if row_idx == min_row:
                # 這是合併區域的第一行
                phonetic_value = phonetic_cell.value
                phonetic_rowspan = max_row - min_row + 1
                row_data['phonetic'] = str(phonetic_value) if phonetic_value else ''
                row_data['phoneticRowSpan'] = phonetic_rowspan
                if phonetic_rowspan > 1:
                    phonetic_merge_state = {
                        'value': str(phonetic_value) if phonetic_value else '',
                        'remaining': phonetic_rowspan - 1
                    }
            else:
                # 這是合併區域的後續行
                if phonetic_merge_state and phonetic_merge_state['remaining'] > 0:
                    row_data['phonetic'] = phonetic_merge_state['value']
                    row_data['phoneticRowSpan'] = 0
                    phonetic_merge_state['remaining'] -= 1
                else:
                    # 應該不會到這裡，但以防萬一
                    phonetic_value = ws.cell(min_row, min_col).value
                    row_data['phonetic'] = str(phonetic_value) if phonetic_value else ''
                    row_data['phoneticRowSpan'] = 0
        else:
            # 非合併儲存格
            phonetic_value = phonetic_cell.value
            row_data['phonetic'] = str(phonetic_value) if phonetic_value else ''
            row_data['phoneticRowSpan'] = 1

        # Columns D-S (16個方言點)
        for dialect_idx in range(16):
            col_idx = 4 + dialect_idx  # D=4, E=5, ..., S=19
            cell = ws.cell(row_idx, col_idx)

            # 檢查是否在合併區域內
            merged_range = is_cell_merged(ws, row_idx, col_idx)

            if merged_range:
                # 取得合併區域的左上角儲存格值
                # bounds 格式：(min_col, min_row, max_col, max_row)
                min_col, min_row, max_col, max_row = merged_range.bounds

                if row_idx == min_row:
                    # 這是合併區域的第一行，使用值並記錄狀態
                    value = ws.cell(min_row, min_col).value
                    value_str = str(value) if value else '-'
                    rowspan = max_row - min_row + 1

                    if rowspan > 1:
                        dialect_merge_state[col_idx] = {
                            'value': value_str,
                            'remaining': rowspan - 1
                        }
                    row_data['dialectValues'].append(value_str)
                else:
                    # 這是合併區域的後續行，使用記錄的值
                    if col_idx in dialect_merge_state and dialect_merge_state[col_idx]['remaining'] > 0:
                        row_data['dialectValues'].append(dialect_merge_state[col_idx]['value'])
                        dialect_merge_state[col_idx]['remaining'] -= 1
                    else:
                        # 意外情況：應該有記錄但沒有，直接取值
                        value = ws.cell(min_row, min_col).value
                        row_data['dialectValues'].append(str(value) if value else '-')
            else:
                # 非合併儲存格
                value = cell.value
                row_data['dialectValues'].append(str(value) if value else '-')

        # 檢查是否有底部邊框（檢查第一個資料儲存格）
        first_data_cell = ws.cell(row_idx, 1)
        if has_bottom_border(first_data_cell):
            row_data['hasBorder'] = True

        result.append(row_data)

    return result

def main():
    # 更新檔案路徑為 data 目錄
    excel_file = 'data/hinghua_rhymes.xlsx'
    output_file = 'website/src/data/rhymeTableData.json'

    if not os.path.exists(excel_file):
        # 如果 data/ 中沒有，嘗試根目錄
        excel_file = 'hinghua_rhymes.xlsx'
        if not os.path.exists(excel_file):
            print(f"錯誤：找不到 Excel 檔案")
            print(f"請確認檔案位於 data/hinghua_rhymes.xlsx 或 hinghua_rhymes.xlsx")
            sys.exit(1)

    print(f"解析 {excel_file}...")
    data = parse_excel_rhyme_table(excel_file)

    print(f"提取了 {len(data)} 列資料")

    # 建立輸出目錄
    os.makedirs('website/src/data', exist_ok=True)

    # 儲存為 JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"資料已儲存至 {output_file}")

    # 顯示前幾列作為範例
    print("\n前 5 列：")
    for i, row in enumerate(data[:5]):
        print(f"\n第 {i+1} 列:")
        print(f"  韻母字: {row['letter']}")
        print(f"  例字: {row['examples'][:50] if len(row['examples']) > 50 else row['examples']}")
        print(f"  擬音: {row['phonetic']}")
        print(f"  莆田 (D): {row['dialectValues'][0]}")
        print(f"  仙遊 (L): {row['dialectValues'][11]}")
        print(f"  RowSpan: {row['rowSpan']}")
        print(f"  PhoneticRowSpan: {row['phoneticRowSpan']}")
        print(f"  HasBorder: {row['hasBorder']}")

    # 顯示跨列的範例
    print("\n\n跨列資料範例：")
    for i, row in enumerate(data):
        if row['rowSpan'] > 1:
            print(f"\n第 {i+1} 列 (跨 {row['rowSpan']} 列):")
            print(f"  韻母字: {row['letter']}")
            print(f"  例字: {row['examples']}")
            print(f"  擬音: {row['phonetic']}")
            print(f"  PhoneticRowSpan: {row['phoneticRowSpan']}")

            # 顯示下一列
            if i+1 < len(data):
                next_row = data[i+1]
                print(f"  下一列 (rowSpan={next_row['rowSpan']}, phoneticRowSpan={next_row['phoneticRowSpan']}):")
                print(f"    韻母字: {next_row['letter']}")
                print(f"    例字: {next_row['examples']}")
                print(f"    擬音: {next_row['phonetic']}")

if __name__ == '__main__':
    main()
